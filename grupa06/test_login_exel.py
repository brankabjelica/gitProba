import unittest
from selenium import webdriver
import os
from openpyxl import load_workbook
import time


class HerokupLogIn(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Chrome()


    def testLoginExcel(self):
        driver = self.driver
        filePath = os.path.join(os.getcwd(), "data", "login.xlsx")
        print(filePath)

        book = load_workbook(filePath)
        sheet = book["login"]

        username = sheet["B2"].value
        password = sheet["C2"].value
        driver.get("http://the-internet.herokuapp.com/login")
        time.sleep(2)

        us = driver.find_element_by_id("username")
        ps = driver.find_element_by_id("password")
        bt = driver.find_element_by_xpath("//i[contains(.,'Login')]")

        time.sleep(3)
        us.send_keys(username)
        time.sleep(2)
        ps.send_keys(password)
        bt.click()
        time.sleep(4)

        try:
            driver.find_element_by_id("flash")
            sheet.cell(row=2, column=4, value='+')
        except:
            sheet.cell(row=2, column=4, value='-')

        book.save(filename=filePath)

    def tearDown(self):
            self.driver.close()


if __name__ == "__main__":
        unittest.main()